#!/usr/bin/env python3
"""竞品筛选助手原型。

基于竞品汇总 Excel 的汇总页，提供：
1. 全表概览
2. 结构化筛选
3. 中文查询
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path


WORKBOOK_CANDIDATES = (
    "竞品数据汇总表V36.xlsx",
    "竞品数据汇总表0305.xlsx",
)
SUMMARY_SHEET_CANDIDATES = (
    "数据汇总表",
    "00 数据汇总表",
)
HEADER_ROW_SCAN_LIMIT = 20
HEADER_TOKENS = ("分组", "类别", "菌株", "外部编号", "来源", "时间节点")
INTERNAL_CATEGORIES = {"WK", "CK1", "CK2", "CK3", "CK4", "CK5", "CK6", "SC"}
COMPARE_INTENT_KEYWORDS = ("对比", "比较", "vs", "VS", "pk", "PK", "差异", "优劣", "PK")
SUMMARY_INTENT_KEYWORDS = ("概览", "汇总", "统计", "分布", "总览", "总体")
NUMBER_PATTERN = re.compile(r"-?\d+(?:\.\d+)?")
GROUP_PATTERN = re.compile(r"(?:分组|组)\s*([0-9]{1,2})")

BASE_COLUMNS = {
    "group_label": 2,
    "category": 3,
    "strain_name": 4,
    "internal_id": 5,
    "external_name": 6,
    "source": 7,
    "owner": 8,
    "time_node": 9,
    "remark": 10,
}

METRIC_COLUMNS = {
    "cell_count": 11,
    "wgs": 12,
    "smell": 13,
    "color": 14,
    "texture": 15,
    "mouthfeel": 16,
    "acid": 17,
    "bitter": 18,
    "salty": 19,
    "fishy": 20,
    "moisture": 21,
    "water_activity": 22,
    "na": 23,
    "ka": 24,
    "cr": 25,
    "arsenic": 26,
    "cadmium": 27,
    "mercury": 28,
    "lead": 29,
    "protein_pct": 30,
    "ash_pct": 31,
    "lactose_ppm": 32,
    "color_50_initial": 33,
    "color_50_3d": 34,
    "color_50_7d": 35,
    "stability_initial": 36,
    "stability_37_1w": 37,
    "stability_37_2w": 38,
    "stability_37_3w": 39,
    "stability_37_4w": 40,
    "stability_37_6w": 41,
    "stability_37_8w": 42,
    "stability_25_1m": 43,
    "stability_25_2m": 44,
    "stability_25_3m": 45,
    "stability_4_1m": 46,
    "stability_4_3m": 47,
}

DISPLAY_COLUMNS = [
    ("row_number", "行号"),
    ("group_label", "分组"),
    ("category", "类别"),
    ("strain_name", "菌株"),
    ("internal_id", "内部编号"),
    ("external_name", "外部编号"),
    ("source", "来源"),
    ("owner", "对接人"),
    ("time_node", "时间"),
    ("remark", "备注"),
    ("cell_count", "总菌体/活菌"),
    ("wgs", "WGS"),
    ("smell", "气味"),
    ("color", "色泽"),
    ("moisture", "水分"),
    ("lactose_ppm", "乳糖"),
]

NUMERIC_ALIAS_MAP = {
    "活菌数": "cell_count",
    "乳酸菌数": "cell_count",
    "总菌体数": "cell_count",
    "WGS": "wgs",
    "气味": "smell",
    "色泽": "color",
    "组织": "texture",
    "口感": "mouthfeel",
    "酸": "acid",
    "苦": "bitter",
    "咸": "salty",
    "腥": "fishy",
    "水分": "moisture",
    "水活": "water_activity",
    "蛋白": "protein_pct",
    "灰分": "ash_pct",
    "乳糖": "lactose_ppm",
}

FIELD_ALIASES = {
    "category": {
        "竞品": {"竞品"},
        "WK": {"wk"},
        "CK1": {"ck1", "对照1", "对照一"},
        "CK2": {"ck2", "对照2", "对照二"},
        "CK3": {"ck3", "对照3", "对照三"},
        "CK4": {"ck4", "对照4", "对照四"},
        "CK5": {"ck5", "对照5", "对照五"},
        "CK6": {"ck6", "对照6", "对照六"},
        "SC": {"sc", "生产"},
    },
    "strain_name": {
        "嗜黏蛋白阿克曼氏菌": {"阿克曼", "阿克曼菌", "akk"},
        "植物乳植杆菌": {"植物乳杆菌", "plantarum"},
        "短双歧杆菌": {"短双", "b06"},
        "长双歧杆菌婴儿亚种": {"婴儿双歧", "m63", "m631"},
        "长双歧杆菌长亚种": {"长双"},
        "动物双歧杆菌乳亚种": {"动物双歧", "bb12", "hn019", "bi07"},
        "副干酪乳酪杆菌": {"副干酪"},
        "干酪乳酪杆菌": {"干酪乳酪杆菌", "k1"},
        "鼠李糖乳酪杆菌": {"鼠李糖", "hn001"},
        "嗜酸乳杆菌": {"嗜酸"},
        "罗伊氏粘液乳杆菌": {"罗伊氏"},
        "凝结魏茨曼氏菌": {"bc99"},
    },
}

SPECIAL_CATEGORY_GROUPS = {
    "微康": {"WK", "CK1", "CK2", "CK3", "CK4", "CK5", "CK6", "SC"},
    "内部": {"WK", "CK1", "CK2", "CK3", "CK4", "CK5", "CK6", "SC"},
}

COMPARATOR_ALIASES = {
    ">=": ">=",
    "大于等于": ">=",
    "不少于": ">=",
    "不低于": ">=",
    "<=": "<=",
    "小于等于": "<=",
    "不高于": "<=",
    ">": ">",
    "大于": ">",
    "高于": ">",
    "<": "<",
    "小于": "<",
    "低于": "<",
    "=": "=",
    "等于": "=",
}

COMPARISON_PATTERN = re.compile(
    r"(活菌数|乳酸菌数|总菌体数|WGS|气味|色泽|组织|口感|酸|苦|咸|腥|水分|水活|蛋白|灰分|乳糖)"
    r"\s*(>=|<=|>|<|=|大于等于|小于等于|不少于|不低于|不高于|大于|小于|高于|低于|等于)\s*([0-9]+(?:\.[0-9]+)?)",
    re.IGNORECASE,
)

MONTH_PATTERN = re.compile(r"(20[0-9]{2})[年/-]?([0-9]{1,2})月?")
DATE_PATTERN = re.compile(r"(20[0-9]{2})[年/-]?([0-9]{1,2})(?:[月/-]?([0-9]{1,2})日?)?")


@dataclass
class NumericCondition:
    field_name: str
    comparator: str
    expected: float


@dataclass
class ParsedQuery:
    text: str
    field_filters: dict[str, set[str]] = field(default_factory=dict)
    numeric_conditions: list[NumericCondition] = field(default_factory=list)
    date_tokens: list[str] = field(default_factory=list)
    text_tokens: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class MetricDefinition:
    field_name: str
    label: str
    direction: str


COMPARE_METRICS = [
    MetricDefinition("cell_count", "总菌体/活菌", "higher"),
    MetricDefinition("wgs", "WGS", "higher"),
    MetricDefinition("smell", "气味", "higher"),
    MetricDefinition("color", "色泽", "higher"),
    MetricDefinition("texture", "组织", "higher"),
    MetricDefinition("mouthfeel", "口感", "higher"),
    MetricDefinition("moisture", "水分", "lower"),
    MetricDefinition("water_activity", "水活", "lower"),
    MetricDefinition("lactose_ppm", "乳糖", "lower"),
    MetricDefinition("stability_37_1w", "37℃1w稳定性", "higher"),
    MetricDefinition("stability_25_1m", "25℃1m稳定性", "higher"),
    MetricDefinition("stability_4_1m", "4℃1m稳定性", "higher"),
]


def clean_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).replace("\n", " ").strip()
    return text


def simplify(text: str) -> str:
    return re.sub(r"[^0-9a-zA-Z\u4e00-\u9fa5]+", "", text).lower()


def parse_number(value: str) -> float | None:
    value = clean_value(value)
    if not value or value in {"/", "未检出", "待进行", "进行中", "菌粉不足", "菌粉不够"}:
        return None
    match = NUMBER_PATTERN.search(value)
    if match:
        return float(match.group())
    return None


def default_workbook_path() -> Path:
    base_dir = Path(__file__).resolve().parent
    for name in WORKBOOK_CANDIDATES:
        candidate = base_dir / name
        if candidate.exists():
            return candidate

    matches = sorted(base_dir.glob("竞品数据汇总表*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if matches:
        return matches[0]
    return base_dir / WORKBOOK_CANDIDATES[0]


def resolve_summary_sheet_name(sheetnames: list[str]) -> str:
    for name in SUMMARY_SHEET_CANDIDATES:
        if name in sheetnames:
            return name

    for name in sheetnames:
        if "汇总" in name:
            return name

    if not sheetnames:
        raise KeyError("工作簿中没有工作表")
    return sheetnames[0]


def detect_header_row(sheet) -> int:
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=HEADER_ROW_SCAN_LIMIT, values_only=True),
        start=1,
    ):
        joined = " ".join(clean_value(value) for value in row if clean_value(value))
        if sum(token in joined for token in HEADER_TOKENS) >= 5:
            return row_number
    return 7


def load_records(workbook_path: Path) -> list[dict[str, str | float | None]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"未找到文件: {workbook_path}")

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = resolve_summary_sheet_name(workbook.sheetnames)
    sheet = workbook[sheet_name]
    header_row = detect_header_row(sheet)
    records: list[dict[str, str | float | None]] = []
    current_group = ""
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, min_col=2, max_col=47, values_only=True),
        start=header_row + 1,
    ):
        if not any(value not in (None, "") for value in row):
            continue

        row_values = {column - 2: clean_value(value) for column, value in enumerate(row, start=2)}
        group_label = row_values.get(BASE_COLUMNS["group_label"] - 2, "")
        if group_label:
            current_group = group_label
        else:
            group_label = current_group

        record: dict[str, str | float | None] = {
            "row_number": row_number,
            "group_label": group_label,
        }
        for field_name, column in BASE_COLUMNS.items():
            if field_name == "group_label":
                continue
            record[field_name] = row_values.get(column - 2, "")
        for field_name, column in METRIC_COLUMNS.items():
            raw = row_values.get(column - 2, "")
            record[field_name] = raw
            record[f"{field_name}_numeric"] = parse_number(raw)
        if any(clean_value(record[name]) for name in ("category", "strain_name", "internal_id", "external_name", "source", "owner", "time_node", "remark")):
            records.append(record)
    return records


def summarize_records(records: list[dict[str, str | float | None]], workbook_path: Path) -> str:
    categories = Counter(clean_value(record["category"]) for record in records if clean_value(record["category"]))
    strains = Counter(clean_value(record["strain_name"]) for record in records if clean_value(record["strain_name"]))
    sources = Counter(clean_value(record["source"]) for record in records if clean_value(record["source"]))
    owners = Counter(clean_value(record["owner"]) for record in records if clean_value(record["owner"]))

    coverage_fields = [
        ("cell_count", "总菌体/活菌"),
        ("wgs", "WGS"),
        ("smell", "气味"),
        ("moisture", "水分"),
        ("lactose_ppm", "乳糖"),
        ("stability_37_1w", "37℃ 1w"),
    ]

    lines = [
        f"记录数: {len(records)}",
        f"分组数: {len({clean_value(record['group_label']) for record in records if clean_value(record['group_label'])})}",
        f"菌株数: {len({clean_value(record['strain_name']) for record in records if clean_value(record['strain_name'])})}",
        f"来源方数: {len(sources)}",
        f"对接人数: {len(owners)}",
        "",
        "类别分布:",
    ]
    lines.extend(f"  - {name}: {count}" for name, count in categories.most_common())
    lines.append("")
    lines.append("高频菌株:")
    lines.extend(f"  - {name}: {count}" for name, count in strains.most_common(8))
    lines.append("")
    lines.append("高频来源方:")
    lines.extend(f"  - {name}: {count}" for name, count in sources.most_common(8))
    lines.append("")
    lines.append("字段覆盖率:")
    for field_name, label in coverage_fields:
        count = sum(1 for record in records if clean_value(record[field_name]))
        lines.append(f"  - {label}: {count}/{len(records)}")
    return "\n".join(lines)


def build_dynamic_aliases(records: list[dict[str, str | float | None]]) -> dict[str, dict[str, set[str]]]:
    aliases = {field_name: {key: set(values) for key, values in mapping.items()} for field_name, mapping in FIELD_ALIASES.items()}

    def add_dynamic(field_name: str, field_value: str) -> None:
        if not field_value or field_value == "/":
            return
        aliases.setdefault(field_name, {})
        aliases[field_name].setdefault(field_value, set())
        aliases[field_name][field_value].add(field_value)
        simplified = simplify(field_value)
        if simplified and simplified != field_value and len(simplified) >= 3:
            aliases[field_name][field_value].add(simplified)
        chinese_parts = re.findall(r"[\u4e00-\u9fa5]{2,}", field_value)
        for part in chinese_parts:
            aliases[field_name][field_value].add(part)
        english_parts = re.findall(r"[A-Za-z0-9-]{2,}", field_value)
        for part in english_parts:
            if len(part) >= 3:
                aliases[field_name][field_value].add(part.lower())
        trimmed = re.sub(r"[-（(].*$", "", field_value).strip()
        if trimmed and trimmed != field_value:
            aliases[field_name][field_value].add(trimmed)

    for record in records:
        for field_name in ("source", "owner", "external_name", "internal_id"):
            add_dynamic(field_name, clean_value(record[field_name]))
        group_label = clean_value(record["group_label"])
        if group_label and not group_label.isdigit():
            add_dynamic("group_label", group_label)
    return aliases


def detect_field_filters(query: str, aliases: dict[str, dict[str, set[str]]]) -> dict[str, set[str]]:
    query_simple = simplify(query)
    detected: dict[str, set[str]] = {}
    for field_name, field_aliases in aliases.items():
        for canonical_value, alias_set in field_aliases.items():
            all_aliases = {canonical_value, simplify(canonical_value), *alias_set}
            for alias in all_aliases:
                alias_text = clean_value(alias)
                if not alias_text:
                    continue
                if alias_text == "/" or len(simplify(alias_text)) < 2:
                    continue
                if re.search(r"[\u4e00-\u9fa5]", alias_text):
                    matched = alias_text in query
                else:
                    matched = simplify(alias_text) in query_simple
                if matched:
                    detected.setdefault(field_name, set()).add(canonical_value)
                    break
    return detected


def extract_numeric_conditions(query: str) -> list[NumericCondition]:
    conditions: list[NumericCondition] = []
    for match in COMPARISON_PATTERN.finditer(query):
        alias = match.group(1)
        comparator = COMPARATOR_ALIASES[match.group(2)]
        field_name = NUMERIC_ALIAS_MAP[alias.upper() if alias.upper() == "WGS" else alias]
        conditions.append(NumericCondition(field_name=field_name, comparator=comparator, expected=float(match.group(3))))
    return conditions


def extract_date_tokens(query: str) -> list[str]:
    tokens: list[str] = []
    for match in DATE_PATTERN.finditer(query):
        year = match.group(1)
        month = match.group(2).zfill(2)
        day = match.group(3)
        if day:
            tokens.append(f"{year}{month}{day.zfill(2)}")
        else:
            tokens.append(f"{year}{month}")
    return tokens


def extract_group_tokens(query: str) -> list[str]:
    return [match.group(1) for match in GROUP_PATTERN.finditer(query)]


def extract_text_tokens(query: str) -> list[str]:
    stripped = COMPARISON_PATTERN.sub(" ", query)
    stripped = DATE_PATTERN.sub(" ", stripped)
    stripped = re.sub(r"[，,。；;、/]+", " ", stripped)
    parts = [part.strip() for part in stripped.split() if part.strip()]
    stop_words = {
        "找",
        "找出",
        "筛选",
        "查询",
        "看看",
        "列出",
        "显示",
        "哪些",
        "哪个",
        "有哪些",
        "相关",
        "样品",
        "竞品",
        "记录",
        "数据",
    }
    tokens = [part for part in parts if part not in stop_words and len(simplify(part)) >= 2]
    return tokens


def parse_query(query: str, aliases: dict[str, dict[str, set[str]]]) -> ParsedQuery:
    field_filters = detect_field_filters(query, aliases)
    group_tokens = extract_group_tokens(query)
    if group_tokens:
        field_filters.setdefault("group_label", set()).update(group_tokens)
    for keyword, category_values in SPECIAL_CATEGORY_GROUPS.items():
        if keyword in query:
            field_filters.setdefault("category", set()).update(category_values)
    numeric_conditions = extract_numeric_conditions(query)
    date_tokens = extract_date_tokens(query)
    text_tokens = extract_text_tokens(query)
    if field_filters or numeric_conditions or date_tokens:
        text_tokens = []
    return ParsedQuery(
        text=query,
        field_filters=field_filters,
        numeric_conditions=numeric_conditions,
        date_tokens=date_tokens,
        text_tokens=text_tokens,
    )


def matches_numeric(value: float | None, comparator: str, expected: float) -> bool:
    if value is None:
        return False
    if comparator == ">=":
        return value >= expected
    if comparator == "<=":
        return value <= expected
    if comparator == ">":
        return value > expected
    if comparator == "<":
        return value < expected
    if comparator == "=":
        return value == expected
    raise ValueError(f"不支持的比较符: {comparator}")


def record_text(record: dict[str, str | float | None]) -> str:
    fields = [
        "group_label",
        "category",
        "strain_name",
        "internal_id",
        "external_name",
        "source",
        "owner",
        "time_node",
        "remark",
    ]
    return " ".join(clean_value(record[field_name]) for field_name in fields)


def filter_records(records: list[dict[str, str | float | None]], parsed_query: ParsedQuery) -> list[dict[str, str | float | None]]:
    filtered = records

    for field_name, values in parsed_query.field_filters.items():
        filtered = [record for record in filtered if clean_value(record.get(field_name, "")) in values]

    for condition in parsed_query.numeric_conditions:
        numeric_field = f"{condition.field_name}_numeric"
        filtered = [
            record
            for record in filtered
            if matches_numeric(record.get(numeric_field), condition.comparator, condition.expected)  # type: ignore[arg-type]
        ]

    for token in parsed_query.date_tokens:
        filtered = [record for record in filtered if token in re.sub(r"[^0-9]", "", clean_value(record["time_node"]))]

    for token in parsed_query.text_tokens:
        token_simple = simplify(token)
        filtered = [record for record in filtered if token_simple in simplify(record_text(record))]

    return filtered


def top_counts(records: list[dict[str, str | float | None]], field_name: str, limit: int = 5) -> list[tuple[str, int]]:
    counter = Counter(clean_value(record[field_name]) for record in records if clean_value(record[field_name]))
    return counter.most_common(limit)


def truncate(text: str, width: int) -> str:
    if len(text) <= width:
        return text
    if width <= 1:
        return text[:width]
    return text[: width - 1] + "…"


def format_table(records: list[dict[str, str | float | None]], limit: int) -> str:
    visible = records[:limit]
    widths = {}
    for field_name, label in DISPLAY_COLUMNS:
        max_width = len(label)
        for record in visible:
            max_width = max(max_width, len(clean_value(record[field_name])))
        widths[field_name] = min(max_width, 18)

    header = " | ".join(truncate(label, widths[field_name]).ljust(widths[field_name]) for field_name, label in DISPLAY_COLUMNS)
    divider = "-+-".join("-" * widths[field_name] for field_name, _ in DISPLAY_COLUMNS)
    lines = [header, divider]

    for record in visible:
        line = " | ".join(
            truncate(clean_value(record[field_name]), widths[field_name]).ljust(widths[field_name])
            for field_name, _ in DISPLAY_COLUMNS
        )
        lines.append(line)
    return "\n".join(lines)


def describe_query(parsed_query: ParsedQuery) -> str:
    parts: list[str] = []
    for field_name, values in parsed_query.field_filters.items():
        parts.append(f"{field_name}={','.join(sorted(values))}")
    for condition in parsed_query.numeric_conditions:
        parts.append(f"{condition.field_name}{condition.comparator}{condition.expected:g}")
    for token in parsed_query.date_tokens:
        parts.append(f"time~{token}")
    if parsed_query.text_tokens:
        parts.append(f"text={','.join(parsed_query.text_tokens)}")
    return "；".join(parts) if parts else "未识别到结构化条件，按全文匹配处理"


def has_query_constraints(parsed_query: ParsedQuery) -> bool:
    return bool(parsed_query.field_filters or parsed_query.numeric_conditions or parsed_query.date_tokens or parsed_query.text_tokens)


def group_sort_key(value: str) -> tuple[int, object]:
    value = clean_value(value)
    if value.isdigit():
        return (0, int(value))
    return (1, value)


def format_number(value: float | None) -> str:
    if value is None:
        return "/"
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def meaningful_value(value: object) -> str:
    text = clean_value(value)
    return "" if text == "/" else text


def is_internal_record(record: dict[str, str | float | None]) -> bool:
    category = clean_value(record["category"])
    source = clean_value(record["source"]).upper()
    return category in INTERNAL_CATEGORIES or source == "WK"


def is_competitor_record(record: dict[str, str | float | None]) -> bool:
    return not is_internal_record(record) and bool(clean_value(record["strain_name"]))


def build_query_from_args(args: argparse.Namespace) -> ParsedQuery:
    query = ParsedQuery(text=getattr(args, "query", "") or "")

    if getattr(args, "strain", None):
        query.field_filters["strain_name"] = {args.strain}
    if getattr(args, "source", None):
        query.field_filters["source"] = {args.source}
    if getattr(args, "owner", None):
        query.field_filters["owner"] = {args.owner}
    if getattr(args, "category", None):
        query.field_filters["category"] = {args.category}
    if getattr(args, "group", None):
        query.field_filters["group_label"] = {args.group}
    if getattr(args, "external_name", None):
        query.field_filters["external_name"] = {args.external_name}
    if getattr(args, "date", None):
        query.date_tokens = [re.sub(r"[^0-9]", "", args.date)]
    if getattr(args, "keyword", None):
        query.text_tokens = [args.keyword]

    if getattr(args, "metric", None) and getattr(args, "op", None) and getattr(args, "value", None) is not None:
        metric = NUMERIC_ALIAS_MAP.get(args.metric, args.metric)
        query.numeric_conditions.append(
            NumericCondition(metric, COMPARATOR_ALIASES.get(args.op, args.op), args.value)
        )
    return query


def summarize_filter_result(records: list[dict[str, str | float | None]]) -> list[str]:
    lines: list[str] = []
    for field_name, label in (("strain_name", "菌株"), ("source", "来源"), ("owner", "对接人"), ("category", "类别")):
        counts = top_counts(records, field_name, limit=5)
        if counts:
            formatted = "，".join(f"{name}({count})" for name, count in counts)
            lines.append(f"- {label}: {formatted}")
    return lines


def detect_intent(query: str) -> str:
    query_lower = query.lower()
    if any(keyword.lower() in query_lower for keyword in COMPARE_INTENT_KEYWORDS):
        return "compare"
    if "竞品" in query and any(keyword in query for keyword in ("微康", "内部", "WK", "CK")):
        return "compare"
    if any(keyword in query for keyword in SUMMARY_INTENT_KEYWORDS):
        return "summary"
    return "filter"


def group_records(records: list[dict[str, str | float | None]]) -> dict[str, list[dict[str, str | float | None]]]:
    grouped: dict[str, list[dict[str, str | float | None]]] = {}
    for record in records:
        group_label = clean_value(record["group_label"])
        grouped.setdefault(group_label, []).append(record)
    for group_label in grouped:
        grouped[group_label] = sorted(grouped[group_label], key=lambda item: int(item["row_number"]))  # type: ignore[arg-type]
    return grouped


def record_label(record: dict[str, str | float | None]) -> str:
    name = meaningful_value(record["external_name"]) or meaningful_value(record["internal_id"]) or meaningful_value(record["strain_name"]) or "未命名样本"
    category = meaningful_value(record["category"]) or "-"
    source = meaningful_value(record["source"]) or "-"
    return f"row{clean_value(record['row_number'])} {category}/{source}/{name}"


def raw_snapshot(record: dict[str, str | float | None]) -> str:
    fields = [("cell_count", "总菌体/活菌"), ("wgs", "WGS"), ("smell", "气味"), ("color", "色泽"), ("moisture", "水分"), ("lactose_ppm", "乳糖")]
    parts = [f"{label}={meaningful_value(record[field_name])}" for field_name, label in fields if meaningful_value(record[field_name])]
    return "；".join(parts[:5]) if parts else "该条记录可用数值很少"


def primary_internal_records(internal_records: list[dict[str, str | float | None]]) -> list[dict[str, str | float | None]]:
    wk_records = [record for record in internal_records if clean_value(record["category"]) == "WK"]
    return wk_records if wk_records else internal_records


def metric_is_better(candidate: float, reference: float, direction: str) -> bool:
    if direction == "higher":
        return candidate > reference
    return candidate < reference


def metric_score_against_reference(candidate: float, reference: float, direction: str) -> int:
    if abs(candidate - reference) < 1e-9:
        return 0
    return 1 if metric_is_better(candidate, reference, direction) else -1


def compare_metric_to_internal(
    competitor: dict[str, str | float | None],
    primary_refs: list[dict[str, str | float | None]],
    all_internal_refs: list[dict[str, str | float | None]],
    metric: MetricDefinition,
) -> tuple[int, str] | None:
    competitor_raw = clean_value(competitor[metric.field_name])
    competitor_num = competitor.get(f"{metric.field_name}_numeric")
    if not competitor_raw or competitor_num is None:
        return None

    primary_numeric_refs = [
        record for record in primary_refs if record.get(f"{metric.field_name}_numeric") is not None
    ]
    all_numeric_refs = [
        record for record in all_internal_refs if record.get(f"{metric.field_name}_numeric") is not None
    ]
    if not primary_numeric_refs and not all_numeric_refs:
        return None

    active_refs = primary_numeric_refs or all_numeric_refs
    values = [record[f"{metric.field_name}_numeric"] for record in active_refs]  # type: ignore[index]
    values = [value for value in values if value is not None]
    if not values:
        return None

    if len(active_refs) == 1:
        ref_record = active_refs[0]
        ref_raw = clean_value(ref_record[metric.field_name])
        ref_num = ref_record[f"{metric.field_name}_numeric"]  # type: ignore[index]
        assert isinstance(ref_num, float)
        score = metric_score_against_reference(competitor_num, ref_num, metric.direction)  # type: ignore[arg-type]
        if score == 0:
            verdict = "与内部持平"
        elif score > 0:
            verdict = "竞品值更优" if metric.direction == "lower" else "竞品更高"
        else:
            verdict = "内部值更优" if metric.direction == "lower" else "内部更高"
        line = f"{metric.label}: {competitor_raw} vs {record_label(ref_record)}={ref_raw}，{verdict}"
        if len(all_numeric_refs) > 1:
            better_count = sum(metric_is_better(competitor_num, ref[f'{metric.field_name}_numeric'], metric.direction) for ref in all_numeric_refs)  # type: ignore[index,arg-type]
            line += f"；在 {len(all_numeric_refs)} 个内部样本中优于 {better_count} 个"
        return score, line

    min_value = min(values)
    max_value = max(values)
    if metric.direction == "higher":
        if competitor_num > max_value:  # type: ignore[operator]
            score = 1
            position = "高于全部内部样本"
        elif competitor_num < min_value:  # type: ignore[operator]
            score = -1
            position = "低于全部内部样本"
        else:
            score = 0
            position = "位于内部范围内"
        better_count = sum(competitor_num > value for value in values)  # type: ignore[operator]
        relation = f"高于 {better_count}/{len(values)} 个内部样本"
    else:
        if competitor_num < min_value:  # type: ignore[operator]
            score = 1
            position = "低于全部内部样本"
        elif competitor_num > max_value:  # type: ignore[operator]
            score = -1
            position = "高于全部内部样本"
        else:
            score = 0
            position = "位于内部范围内"
        better_count = sum(competitor_num < value for value in values)  # type: ignore[operator]
        relation = f"低于 {better_count}/{len(values)} 个内部样本"
    line = (
        f"{metric.label}: {competitor_raw}；内部范围 {format_number(min_value)}-{format_number(max_value)}，"
        f"{position}，{relation}"
    )
    return score, line


def compare_competitor_record(
    competitor: dict[str, str | float | None],
    internal_records: list[dict[str, str | float | None]],
    max_metrics: int,
) -> list[str]:
    lines = [f"  对比对象: {record_label(competitor)}"]
    primary_refs = primary_internal_records(internal_records)
    metric_lines: list[str] = []
    better_metrics: list[str] = []
    lower_metrics: list[str] = []
    neutral_metrics: list[str] = []

    for metric in COMPARE_METRICS:
        result = compare_metric_to_internal(competitor, primary_refs, internal_records, metric)
        if result is None:
            continue
        score, line = result
        metric_lines.append(f"    - {line}")
        if score > 0:
            better_metrics.append(metric.label)
        elif score < 0:
            lower_metrics.append(metric.label)
        else:
            neutral_metrics.append(metric.label)
        if len(metric_lines) >= max_metrics:
            break

    if not metric_lines:
        lines.append(f"    - 可比数值不足。原始值: {raw_snapshot(competitor)}")
        return lines

    lines.extend(metric_lines)
    verdict_parts: list[str] = []
    if better_metrics:
        verdict_parts.append(f"竞品占优={ '、'.join(better_metrics[:4]) }")
    if lower_metrics:
        verdict_parts.append(f"内部占优={ '、'.join(lower_metrics[:4]) }")
    if neutral_metrics and not better_metrics and not lower_metrics:
        verdict_parts.append("主要指标位于内部范围内")
    if len(metric_lines) < min(max_metrics, len(COMPARE_METRICS)):
        verdict_parts.append("其余指标缺失或量表不统一")
    lines.append(f"    初步判断: {'；'.join(verdict_parts) if verdict_parts else '信息不足'}")
    return lines


def rank_competitors_without_internal(competitors: list[dict[str, str | float | None]]) -> str:
    numeric_competitors = [record for record in competitors if record.get("cell_count_numeric") is not None]
    if not numeric_competitors:
        return "当前分组没有 WK/CK/SC 参考，且竞品总菌体/活菌字段也缺少可比数值。"
    ranked = sorted(
        numeric_competitors,
        key=lambda record: record["cell_count_numeric"],  # type: ignore[index]
        reverse=True,
    )[:5]
    summary = "；".join(
        f"{record_label(record)}={clean_value(record['cell_count'])}" for record in ranked
    )
    return f"当前分组没有 WK/CK/SC 参考，无法做自动对比。按总菌体/活菌排序: {summary}"


def render_compare_group(
    group_label: str,
    group_records_list: list[dict[str, str | float | None]],
    focus_competitors: list[dict[str, str | float | None]],
    max_competitors: int,
    max_metrics: int,
) -> list[str]:
    strains = sorted({clean_value(record["strain_name"]) for record in group_records_list if clean_value(record["strain_name"])})
    strain_text = " / ".join(strains[:3]) if strains else "未识别菌株"
    if len(strains) > 3:
        strain_text += " / ..."

    internal_records = [record for record in group_records_list if is_internal_record(record)]
    competitor_records = [record for record in group_records_list if is_competitor_record(record)]

    lines = [f"分组 {group_label} | {strain_text}"]
    if internal_records:
        lines.append("  内部参考: " + "；".join(record_label(record) for record in internal_records[:6]))
    else:
        lines.append("  内部参考: 无")
    if competitor_records:
        lines.append("  竞品样本: " + "；".join(record_label(record) for record in competitor_records[:6]))
    else:
        lines.append("  竞品样本: 无")

    if not internal_records:
        lines.append(f"  自动结论: {rank_competitors_without_internal(competitor_records)}")
        return lines

    target_competitors = focus_competitors or competitor_records
    if not target_competitors:
        lines.append("  自动结论: 当前筛选条件只命中了内部参考，没有额外竞品可展开。")
        return lines

    for competitor in target_competitors[:max_competitors]:
        lines.extend(compare_competitor_record(competitor, internal_records, max_metrics))
    if len(target_competitors) > max_competitors:
        lines.append(f"  ... 还有 {len(target_competitors) - max_competitors} 个竞品未展开")
    return lines


def build_compare_report(
    records: list[dict[str, str | float | None]],
    parsed_query: ParsedQuery,
    max_groups: int = 5,
    max_competitors: int = 4,
    max_metrics: int = 6,
) -> str:
    grouped = group_records(records)
    matched_records = filter_records(records, parsed_query) if has_query_constraints(parsed_query) else records
    target_groups = sorted(
        {clean_value(record["group_label"]) for record in matched_records if clean_value(record["group_label"])},
        key=group_sort_key,
    )
    if not target_groups:
        return "没有找到可对比的分组。"

    lines = []
    if has_query_constraints(parsed_query):
        lines.append(f"解析: {describe_query(parsed_query)}")
    lines.append(f"命中分组: {len(target_groups)}")
    if len(target_groups) > max_groups:
        lines.append(f"展示前 {max_groups} 个分组，其余 {len(target_groups) - max_groups} 个分组已省略。")
    lines.append("")

    for group_label in target_groups[:max_groups]:
        group_records_list = grouped[group_label]
        focus_competitors = [
            record for record in matched_records
            if clean_value(record["group_label"]) == group_label and is_competitor_record(record)
        ]
        lines.extend(render_compare_group(group_label, group_records_list, focus_competitors, max_competitors, max_metrics))
        lines.append("")
    return "\n".join(lines).rstrip()


def execute_filter(records: list[dict[str, str | float | None]], parsed_query: ParsedQuery, limit: int) -> str:
    filtered = filter_records(records, parsed_query)
    lines = [
        f"解析: {describe_query(parsed_query)}",
        f"命中: {len(filtered)} 条",
        "",
    ]
    if not filtered:
        lines.append("没有找到匹配记录。")
        return "\n".join(lines)

    lines.append(format_table(filtered, limit))
    summary_lines = summarize_filter_result(filtered)
    if summary_lines:
        lines.append("")
        lines.append("结果摘要:")
        lines.extend(summary_lines)
    return "\n".join(lines)


def execute_agent(
    records: list[dict[str, str | float | None]],
    workbook_path: Path,
    query: str,
    limit: int,
    max_groups: int,
    max_competitors: int,
    max_metrics: int,
) -> str:
    intent = detect_intent(query)
    aliases = build_dynamic_aliases(records)
    parsed = parse_query(query, aliases)
    lines = [f"问题: {query}", f"任务识别: {intent}", ""]
    if intent == "summary":
        lines.append(summarize_records(records, workbook_path))
        return "\n".join(lines)
    if intent == "compare":
        lines.append(build_compare_report(records, parsed, max_groups=max_groups, max_competitors=max_competitors, max_metrics=max_metrics))
        return "\n".join(lines)
    lines.append(execute_filter(records, parsed, limit))
    return "\n".join(lines)


def run_summary(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook).expanduser().resolve()
    records = load_records(workbook_path)
    print(summarize_records(records, workbook_path))
    return 0


def run_list(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook).expanduser().resolve()
    records = load_records(workbook_path)
    query = build_query_from_args(args)
    print(execute_filter(records, query, args.limit))
    return 0


def run_compare(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook).expanduser().resolve()
    records = load_records(workbook_path)
    if args.query:
        aliases = build_dynamic_aliases(records)
        parsed = parse_query(args.query, aliases)
    else:
        parsed = build_query_from_args(args)
    print(build_compare_report(records, parsed, max_groups=args.max_groups, max_competitors=args.max_competitors, max_metrics=args.max_metrics))
    return 0


def run_ask(args: argparse.Namespace) -> int:
    workbook_path = Path(args.workbook).expanduser().resolve()
    records = load_records(workbook_path)
    print(
        execute_agent(
            records,
            workbook_path,
            args.query,
            limit=args.limit,
            max_groups=args.max_groups,
            max_competitors=args.max_competitors,
            max_metrics=args.max_metrics,
        )
    )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="竞品智能体原型")
    parser.add_argument(
        "--workbook",
        default=str(default_workbook_path()),
        help="Excel 文件路径，默认优先读取脚本同目录下最新的竞品数据汇总表",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    summary_parser = subparsers.add_parser("summary", help="查看全表概览")
    summary_parser.set_defaults(func=run_summary)

    list_parser = subparsers.add_parser("list", help="按结构化条件筛选")
    list_parser.add_argument("--strain", help="菌株名称")
    list_parser.add_argument("--source", help="来源方")
    list_parser.add_argument("--owner", help="对接人")
    list_parser.add_argument("--category", help="类别，如竞品/WK/CK1")
    list_parser.add_argument("--group", help="分组编号或分组名称")
    list_parser.add_argument("--external-name", dest="external_name", help="外部编号/名称，如 HN019")
    list_parser.add_argument("--date", help="时间节点，可传 20260225 或 2026-02")
    list_parser.add_argument("--keyword", help="额外关键词，会做全文包含匹配")
    list_parser.add_argument("--metric", help="数值字段，如 活菌数/WGS/水分/乳糖")
    list_parser.add_argument("--op", help="比较符，如 >= <= > < =")
    list_parser.add_argument("--value", type=float, help="数值阈值")
    list_parser.add_argument("--limit", type=int, default=20, help="最多显示多少条")
    list_parser.set_defaults(func=run_list)

    compare_parser = subparsers.add_parser("compare", help="竞品 vs WK/CK 自动对比")
    compare_parser.add_argument("--query", help="自然语言对比请求，例如：对比阿克曼菌竞品和微康")
    compare_parser.add_argument("--strain", help="菌株名称")
    compare_parser.add_argument("--source", help="来源方")
    compare_parser.add_argument("--owner", help="对接人")
    compare_parser.add_argument("--category", help="类别，如竞品/WK/CK1")
    compare_parser.add_argument("--group", help="分组编号或分组名称")
    compare_parser.add_argument("--external-name", dest="external_name", help="外部编号/名称，如 HN019")
    compare_parser.add_argument("--date", help="时间节点，可传 20260225 或 2026-02")
    compare_parser.add_argument("--keyword", help="额外关键词，会做全文包含匹配")
    compare_parser.add_argument("--metric", help="数值字段，如 活菌数/WGS/水分/乳糖")
    compare_parser.add_argument("--op", help="比较符，如 >= <= > < =")
    compare_parser.add_argument("--value", type=float, help="数值阈值")
    compare_parser.add_argument("--max-groups", type=int, default=5, help="最多展开多少个分组")
    compare_parser.add_argument("--max-competitors", type=int, default=4, help="每个分组最多展开多少个竞品")
    compare_parser.add_argument("--max-metrics", type=int, default=6, help="每个竞品最多展示多少个可比指标")
    compare_parser.set_defaults(func=run_compare)

    ask_parser = subparsers.add_parser("ask", help="智能体入口，自动识别筛选/对比/概览")
    ask_parser.add_argument("query", help="例如：找出阿克曼菌，来源善恩康，2026年2月送检的竞品")
    ask_parser.add_argument("--limit", type=int, default=20, help="最多显示多少条")
    ask_parser.add_argument("--max-groups", type=int, default=5, help="对比任务最多展开多少个分组")
    ask_parser.add_argument("--max-competitors", type=int, default=4, help="对比任务每组最多展开多少个竞品")
    ask_parser.add_argument("--max-metrics", type=int, default=6, help="对比任务每个竞品最多展示多少个指标")
    ask_parser.set_defaults(func=run_ask)

    agent_parser = subparsers.add_parser("agent", help="ask 的别名，作为智能体入口")
    agent_parser.add_argument("query", help="例如：对比植物乳植杆菌竞品和微康")
    agent_parser.add_argument("--limit", type=int, default=20, help="筛选任务最多显示多少条")
    agent_parser.add_argument("--max-groups", type=int, default=5, help="对比任务最多展开多少个分组")
    agent_parser.add_argument("--max-competitors", type=int, default=4, help="对比任务每组最多展开多少个竞品")
    agent_parser.add_argument("--max-metrics", type=int, default=6, help="对比任务每个竞品最多展示多少个指标")
    agent_parser.set_defaults(func=run_ask)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
